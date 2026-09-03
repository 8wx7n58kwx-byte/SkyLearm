"""
EdTech RBAC prototype backend.
FastAPI + SQLite. No external auth deps: passwords hashed with PBKDF2
(hashlib, stdlib) and session tokens are a minimal HMAC-signed payload
(stdlib hmac + base64) -- same "no extra dependency" philosophy as the
existing course-hub api_server.py (which uses hmac.compare_digest for its
admin key).
"""
import base64
import hashlib
import hmac
import json
import os
import random
import secrets
import sqlite3
import time
from pathlib import Path
from typing import Optional

from typing import Any, Dict, List

from io import BytesIO

from fastapi import FastAPI, HTTPException, Header, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

QUIZ_TYPES = ("multiple_choice", "short_answer", "file_upload", "numeric", "sequencing", "hotspot")
UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DB_PATH = Path(__file__).parent / "data.db"
SECRET_KEY_PATH = Path(__file__).parent / ".session_secret"


def _load_or_create_secret_key() -> bytes:
    """Resolve the HMAC signing key for session tokens.

    Never falls back to a hardcoded/well-known string -- that would let
    anyone forge a valid session token (including an admin token) for this
    publicly reachable prototype. Prefer an explicit SECRET_KEY env var;
    otherwise generate a random key on first run and persist it to a
    gitignored local file so tokens issued by this process stay valid across
    restarts (a fresh deploy/reset gets a fresh, unguessable key).
    """
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key.encode()
    if SECRET_KEY_PATH.exists():
        return SECRET_KEY_PATH.read_bytes()
    key = secrets.token_bytes(32)
    SECRET_KEY_PATH.write_bytes(key)
    return key


SECRET_KEY = _load_or_create_secret_key()
TOKEN_TTL_SECONDS = 8 * 3600

app = FastAPI(title="EdTech RBAC Prototype")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


# ---------- password hashing (PBKDF2, stdlib only) ----------
def hash_password(password: str, salt: Optional[bytes] = None) -> str:
    salt = salt or secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200_000)
    return base64.b64encode(salt).decode() + "$" + base64.b64encode(dk).decode()


def verify_password(password: str, stored: str) -> bool:
    try:
        salt_b64, dk_b64 = stored.split("$")
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(dk_b64)
    except Exception:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200_000)
    return hmac.compare_digest(dk, expected)


# ---------- minimal signed session token (stdlib only) ----------
def make_token(user_id: int, role: str) -> str:
    payload = {"uid": user_id, "role": role, "exp": int(time.time()) + TOKEN_TTL_SECONDS}
    raw = json.dumps(payload).encode()
    body = base64.urlsafe_b64encode(raw).decode()
    sig = hmac.new(SECRET_KEY, body.encode(), hashlib.sha256).hexdigest()
    return f"{body}.{sig}"


def read_token(token: str) -> dict:
    try:
        body, sig = token.split(".")
        expected_sig = hmac.new(SECRET_KEY, body.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            raise ValueError("bad signature")
        payload = json.loads(base64.urlsafe_b64decode(body))
        if payload["exp"] < time.time():
            raise ValueError("expired")
        return payload
    except Exception:
        raise HTTPException(401, "Invalid or expired session")


# ---------- db ----------
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = db()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','instructor','student')),
            student_id TEXT UNIQUE,
            created_by INTEGER REFERENCES users(id),
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            course_url TEXT,
            instructor_id INTEGER NOT NULL REFERENCES users(id),
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS enrollments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            course_id INTEGER NOT NULL REFERENCES courses(id),
            granted_by INTEGER NOT NULL REFERENCES users(id),
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(user_id, course_id)
        );
        CREATE TABLE IF NOT EXISTS chapters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id INTEGER NOT NULL REFERENCES courses(id),
            title TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chapter_id INTEGER NOT NULL REFERENCES chapters(id),
            type TEXT NOT NULL CHECK(type IN ('text','image','audio')),
            title TEXT NOT NULL,
            body TEXT,
            file_url TEXT,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS quizzes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chapter_id INTEGER NOT NULL REFERENCES chapters(id),
            title TEXT NOT NULL,
            instructions TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quiz_id INTEGER NOT NULL REFERENCES quizzes(id),
            quiz_type TEXT NOT NULL CHECK(quiz_type IN
                ('multiple_choice','short_answer','file_upload','numeric','sequencing','hotspot')),
            prompt TEXT NOT NULL,
            points REAL NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            params_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question_id INTEGER NOT NULL REFERENCES questions(id),
            student_id INTEGER NOT NULL REFERENCES users(id),
            answer_json TEXT NOT NULL,
            is_correct INTEGER,
            score REAL,
            grading_state TEXT NOT NULL DEFAULT 'auto' CHECK(grading_state IN ('auto','pending','graded')),
            feedback TEXT,
            submitted_at TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(question_id, student_id)
        );
        """
    )
    conn.commit()

    # seed only if empty
    if conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] == 0:
        admin_id = conn.execute(
            "INSERT INTO users (email,password_hash,full_name,role) VALUES (?,?,?,?)",
            ("admin@edtech.local", hash_password("admin123"), "Admin One", "admin"),
        ).lastrowid
        instr_id = conn.execute(
            "INSERT INTO users (email,password_hash,full_name,role,created_by) VALUES (?,?,?,?,?)",
            ("instructor@edtech.local", hash_password("instructor123"), "Instructor One", "instructor", admin_id),
        ).lastrowid
        student_id = conn.execute(
            "INSERT INTO users (email,password_hash,full_name,role,student_id,created_by) VALUES (?,?,?,?,?,?)",
            ("student@edtech.local", hash_password("student123"), "Student Demo", "student", "68010001", instr_id),
        ).lastrowid
        bendlab_id = conn.execute(
            "INSERT INTO courses (slug,title,description,course_url,instructor_id) VALUES (?,?,?,?,?)",
            ("bendlab", "BendLab — พับโลหะแผ่น", "K-Factor, Bend Allowance, Setback, Y-Factor",
             "https://bendlab-course.pplx.app/bendlab/index.html", instr_id),
        ).lastrowid
        conn.execute(
            "INSERT INTO courses (slug,title,description,course_url,instructor_id) VALUES (?,?,?,?,?)",
            ("gearlab", "GearLab — เกียร์", "Module, Involute Profile, Center Distance, Backlash",
             "https://bendlab-course.pplx.app/gear/index.html", instr_id),
        )
        # demo: student is whitelisted into bendlab only, NOT gearlab --
        # proves whitelisting actually gates access per course.
        conn.execute(
            "INSERT INTO enrollments (user_id,course_id,granted_by) VALUES (?,?,?)",
            (student_id, bendlab_id, instr_id),
        )
        conn.commit()
    conn.close()


init_db()


def heal_locked_out_admins():
    """Safety net: if every admin account is inactive (e.g. an admin accidentally
    deactivated their own account and there is no other admin left to fix it),
    reactivate the oldest admin account on startup so the system never gets
    permanently locked out."""
    conn = db()
    active_admins = conn.execute("SELECT COUNT(*) c FROM users WHERE role='admin' AND is_active=1").fetchone()["c"]
    if active_admins == 0:
        row = conn.execute("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone()
        if row:
            conn.execute("UPDATE users SET is_active=1 WHERE id=?", (row["id"],))
            conn.commit()
    conn.close()


heal_locked_out_admins()


# ---------- auth dependency ----------
def current_user(authorization: str = Header(default="")) -> sqlite3.Row:
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing bearer token")
    payload = read_token(authorization.removeprefix("Bearer ").strip())
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE id=? AND is_active=1", (payload["uid"],)).fetchone()
    conn.close()
    if not user:
        raise HTTPException(401, "Account not found or deactivated")
    return user


def require_role(*roles):
    def _dep(user: sqlite3.Row = Depends(current_user)) -> sqlite3.Row:
        if user["role"] not in roles:
            raise HTTPException(403, "Insufficient permission for this role")
        return user

    return _dep


def user_public(u: sqlite3.Row) -> dict:
    return {
        "id": u["id"], "email": u["email"], "full_name": u["full_name"], "role": u["role"],
        "student_id": u["student_id"], "is_active": bool(u["is_active"]), "created_at": u["created_at"],
    }


# ================= AUTH =================
class LoginBody(BaseModel):
    email: str
    password: str


@app.post("/api/auth/login")
def login(body: LoginBody):
    conn = db()
    u = conn.execute("SELECT * FROM users WHERE email=?", (body.email,)).fetchone()
    conn.close()
    if not u or not u["is_active"] or not verify_password(body.password, u["password_hash"]):
        raise HTTPException(401, "อีเมลหรือรหัสผ่านไม่ถูกต้อง")
    token = make_token(u["id"], u["role"])
    return {"token": token, "role": u["role"], "full_name": u["full_name"], "redirect_to": f"/{u['role']}"}


@app.get("/api/auth/me")
def me(user: sqlite3.Row = Depends(current_user)):
    return user_public(user)


@app.get("/api/health")
def health():
    return {"ok": True}


# ================= ADMIN: full user CRUD =================
class CreateUserBody(BaseModel):
    email: str
    password: str
    full_name: str
    role: str
    student_id: Optional[str] = None


class UpdateUserBody(BaseModel):
    full_name: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    role: Optional[str] = None


@app.get("/api/admin/users")
def admin_list_users(admin: sqlite3.Row = Depends(require_role("admin"))):
    conn = db()
    rows = conn.execute("SELECT * FROM users ORDER BY role, full_name").fetchall()
    conn.close()
    return [user_public(r) for r in rows]


@app.post("/api/admin/users")
def admin_create_user(body: CreateUserBody, admin: sqlite3.Row = Depends(require_role("admin"))):
    if body.role not in ("admin", "instructor", "student"):
        raise HTTPException(400, "role must be admin, instructor, or student")
    conn = db()
    try:
        cur = conn.execute(
            "INSERT INTO users (email,password_hash,full_name,role,student_id,created_by) VALUES (?,?,?,?,?,?)",
            (body.email, hash_password(body.password), body.full_name, body.role, body.student_id, admin["id"]),
        )
        conn.commit()
    except sqlite3.IntegrityError as e:
        raise HTTPException(409, f"email or student_id already exists ({e})")
    finally:
        conn.close()
    return {"id": cur.lastrowid}


@app.patch("/api/admin/users/{uid}")
def admin_update_user(uid: int, body: UpdateUserBody, admin: sqlite3.Row = Depends(require_role("admin"))):
    conn = db()
    if body.is_active is False or (body.role is not None and body.role != "admin"):
        target = conn.execute("SELECT role, is_active FROM users WHERE id=?", (uid,)).fetchone()
        if target and target["role"] == "admin" and target["is_active"]:
            other_active_admins = conn.execute(
                "SELECT COUNT(*) c FROM users WHERE role='admin' AND is_active=1 AND id!=?", (uid,)
            ).fetchone()["c"]
            if other_active_admins == 0:
                conn.close()
                raise HTTPException(400, "ไม่สามารถปิดใช้งานหรือเปลี่ยนบทบาทของแอดมินคนสุดท้ายได้ เพื่อป้องกันระบบถูกล็อกออกจากการเข้าสู่การจัดการ")
    fields, values = [], []
    if body.full_name is not None:
        fields.append("full_name=?"); values.append(body.full_name)
    if body.is_active is not None:
        fields.append("is_active=?"); values.append(1 if body.is_active else 0)
    if body.password:
        fields.append("password_hash=?"); values.append(hash_password(body.password))
    if body.role:
        if body.role not in ("admin", "instructor", "student"):
            raise HTTPException(400, "invalid role")
        fields.append("role=?"); values.append(body.role)
    if not fields:
        conn.close()
        raise HTTPException(400, "no fields to update")
    values.append(uid)
    conn.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=?", values)
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/admin/users/{uid}")
def admin_delete_user(uid: int, admin: sqlite3.Row = Depends(require_role("admin"))):
    conn = db()
    target = conn.execute("SELECT role, is_active FROM users WHERE id=?", (uid,)).fetchone()
    if target and target["role"] == "admin" and target["is_active"]:
        other_active_admins = conn.execute(
            "SELECT COUNT(*) c FROM users WHERE role='admin' AND is_active=1 AND id!=?", (uid,)
        ).fetchone()["c"]
        if other_active_admins == 0:
            conn.close()
            raise HTTPException(400, "ไม่สามารถลบแอดมินคนสุดท้ายได้ เพื่อป้องกันระบบถูกล็อกออกจากการเข้าสู่การจัดการ")
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ================= INSTRUCTOR =================
class CreateStudentBody(BaseModel):
    email: str
    password: str
    full_name: str
    student_id: str


class WhitelistBody(BaseModel):
    student_id: str


class CreateCourseBody(BaseModel):
    slug: str
    title: str
    description: Optional[str] = None
    course_url: Optional[str] = None


def _owned_course(conn, course_id: int, instructor_id: int) -> sqlite3.Row:
    c = conn.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
    if not c:
        raise HTTPException(404, "course not found")
    if c["instructor_id"] != instructor_id:
        raise HTTPException(403, "not your course")
    return c


@app.get("/api/instructor/courses")
def instructor_courses(instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    rows = conn.execute("SELECT * FROM courses WHERE instructor_id=?", (instr["id"],)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/instructor/courses")
def instructor_create_course(body: CreateCourseBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    slug = body.slug.strip().lower().replace(" ", "-")
    if not slug or not body.title.strip():
        raise HTTPException(400, "slug and title are required")
    conn = db()
    try:
        cur = conn.execute(
            "INSERT INTO courses (slug,title,description,course_url,instructor_id) VALUES (?,?,?,?,?)",
            (slug, body.title.strip(), body.description, body.course_url, instr["id"]),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(409, "slug นี้ถูกใช้แล้ว กรุณาเปลี่ยน slug")
    finally:
        conn.close()
    return {"id": cur.lastrowid}


@app.post("/api/instructor/students")
def instructor_create_student(body: CreateStudentBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    try:
        cur = conn.execute(
            "INSERT INTO users (email,password_hash,full_name,role,student_id,created_by) VALUES (?,?,?,?,?,?)",
            (body.email, hash_password(body.password), body.full_name, "student", body.student_id, instr["id"]),
        )
        conn.commit()
    except sqlite3.IntegrityError as e:
        raise HTTPException(409, f"email or student_id already exists ({e})")
    finally:
        conn.close()
    return {"id": cur.lastrowid}


@app.get("/api/instructor/courses/{course_id}/students")
def instructor_roster(course_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    rows = conn.execute(
        """SELECT u.id, u.full_name, u.student_id, u.email, e.created_at enrolled_at
           FROM enrollments e JOIN users u ON u.id = e.user_id
           WHERE e.course_id=? ORDER BY u.full_name""",
        (course_id,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/instructor/courses/{course_id}/enrollments")
def instructor_whitelist(course_id: int, body: WhitelistBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    student = conn.execute("SELECT * FROM users WHERE student_id=? AND role='student'", (body.student_id,)).fetchone()
    if not student:
        conn.close()
        raise HTTPException(404, "ไม่พบนักศึกษาที่มีรหัสนี้")
    try:
        conn.execute(
            "INSERT INTO enrollments (user_id,course_id,granted_by) VALUES (?,?,?)",
            (student["id"], course_id, instr["id"]),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(409, "นักศึกษาคนนี้มีสิทธิ์เข้าคอร์สนี้อยู่แล้ว")
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/instructor/courses/{course_id}/enrollments/{uid}")
def instructor_revoke(course_id: int, uid: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    conn.execute("DELETE FROM enrollments WHERE course_id=? AND user_id=?", (course_id, uid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ================= STUDENT (read-only) =================
@app.get("/api/student/courses")
def student_courses(student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    rows = conn.execute(
        """SELECT c.* FROM enrollments e JOIN courses c ON c.id = e.course_id
           WHERE e.user_id=? ORDER BY c.title""",
        (student["id"],),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/student/courses/{course_id}")
def student_course_detail(course_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    row = conn.execute(
        """SELECT c.* FROM enrollments e JOIN courses c ON c.id = e.course_id
           WHERE e.user_id=? AND c.id=?""",
        (student["id"], course_id),
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(403, "คุณไม่มีสิทธิ์เข้าคอร์สนี้ — กรุณาติดต่ออาจารย์ผู้สอน")
    return dict(row)


# ================= FILE UPLOAD (materials, hotspot images, PDF answers) =================
MAX_UPLOAD_MB = 20


@app.post("/api/upload")
def upload_file(file: UploadFile = File(...), user: sqlite3.Row = Depends(current_user)):
    data = file.file.read()
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"ไฟล์ใหญ่เกิน {MAX_UPLOAD_MB}MB")
    ext = Path(file.filename or "").suffix[:10]
    fname = f"{secrets.token_hex(8)}{ext}"
    (UPLOAD_DIR / fname).write_bytes(data)
    return {"url": f"/uploads/{fname}"}


# ================= CHAPTERS / MATERIALS / QUIZZES (instructor authoring) =================
class ChapterBody(BaseModel):
    title: str
    sort_order: Optional[int] = 0


class MaterialBody(BaseModel):
    type: str
    title: str
    body: Optional[str] = None
    file_url: Optional[str] = None
    sort_order: Optional[int] = 0


class QuizBody(BaseModel):
    title: str
    instructions: Optional[str] = None


class QuestionBody(BaseModel):
    quiz_type: str
    prompt: str
    points: Optional[float] = 1
    sort_order: Optional[int] = 0
    params: Dict[str, Any]


class SubmitAnswerBody(BaseModel):
    answer: Dict[str, Any]


class GradeSubmissionBody(BaseModel):
    score: float
    feedback: Optional[str] = None


def _owned_chapter(conn, chapter_id: int, instructor_id: int) -> sqlite3.Row:
    row = conn.execute(
        """SELECT ch.*, c.instructor_id, c.id AS course_id FROM chapters ch
           JOIN courses c ON c.id = ch.course_id WHERE ch.id=?""",
        (chapter_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบบทเรียนนี้")
    if row["instructor_id"] != instructor_id:
        raise HTTPException(403, "not your chapter")
    return row


def _owned_quiz(conn, quiz_id: int, instructor_id: int) -> sqlite3.Row:
    row = conn.execute(
        """SELECT q.*, ch.course_id, c.instructor_id FROM quizzes q
           JOIN chapters ch ON ch.id = q.chapter_id
           JOIN courses c ON c.id = ch.course_id WHERE q.id=?""",
        (quiz_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบควิซนี้")
    if row["instructor_id"] != instructor_id:
        raise HTTPException(403, "not your quiz")
    return row


def _student_owned_chapter(conn, chapter_id: int, student_id: int) -> sqlite3.Row:
    row = conn.execute(
        """SELECT ch.*, c.id AS course_id FROM chapters ch
           JOIN courses c ON c.id = ch.course_id
           JOIN enrollments e ON e.course_id = c.id AND e.user_id = ?
           WHERE ch.id=?""",
        (student_id, chapter_id),
    ).fetchone()
    if not row:
        raise HTTPException(403, "คุณไม่มีสิทธิ์เข้าบทเรียนนี้")
    return row


def _student_owned_quiz(conn, quiz_id: int, student_id: int) -> sqlite3.Row:
    row = conn.execute(
        """SELECT q.*, ch.course_id FROM quizzes q
           JOIN chapters ch ON ch.id = q.chapter_id
           JOIN enrollments e ON e.course_id = ch.course_id AND e.user_id = ?
           WHERE q.id=?""",
        (student_id, quiz_id),
    ).fetchone()
    if not row:
        raise HTTPException(403, "คุณไม่มีสิทธิ์เข้าควิซนี้")
    return row


def grade_answer(quiz_type: str, params: dict, answer: dict) -> Optional[bool]:
    """Returns True/False for auto-graded types, or None for file_upload (manual)."""
    try:
        if quiz_type == "multiple_choice":
            return answer.get("selected_index") == params.get("correct_index")
        if quiz_type == "short_answer":
            text = str(answer.get("text", ""))
            cs = bool(params.get("case_sensitive", False))
            hay = text if cs else text.lower()
            return any((k if cs else str(k).lower()) in hay for k in params.get("keywords") or [])
        if quiz_type == "numeric":
            val = float(answer.get("value"))
            target = float(params.get("target"))
            tol = float(params.get("tolerance", 0))
            return abs(val - target) <= tol
        if quiz_type == "sequencing":
            return list(answer.get("order") or []) == list(params.get("correct_order") or [])
        if quiz_type == "hotspot":
            x, y = answer.get("x"), answer.get("y")
            if x is None or y is None:
                return False
            for z in params.get("zones") or []:
                if z["x"] <= x <= z["x"] + z["w"] and z["y"] <= y <= z["y"] + z["h"]:
                    return True
            return False
    except (TypeError, ValueError, KeyError):
        return False
    return None  # file_upload -> pending manual grading


def strip_answer_key(quiz_type: str, params: dict) -> dict:
    """Remove the correct-answer fields before sending a question to a student."""
    if quiz_type == "multiple_choice":
        return {"options": params.get("options", [])}
    if quiz_type == "sequencing":
        # shuffle display order so the task is non-trivial; each item keeps its
        # original index so the student's rearranged order can be graded against
        # correct_order (set at creation time as the identity [0..n-1]).
        items = params.get("items", [])
        order = list(range(len(items)))
        random.shuffle(order)
        return {"items": [{"idx": i, "text": items[i]} for i in order]}
    if quiz_type == "hotspot":
        return {"image_url": params.get("image_url")}
    if quiz_type == "file_upload":
        return {"accept": params.get("accept"), "max_mb": params.get("max_mb")}
    return {}  # short_answer, numeric: no safe fields to reveal


def letter_grade(percent: Optional[float]) -> Optional[str]:
    if percent is None:
        return None
    if percent >= 90:
        return "A"
    if percent >= 80:
        return "B"
    if percent >= 70:
        return "C"
    if percent >= 60:
        return "D"
    return "F"


def compute_progress(conn, student_id: int, course_id: int) -> dict:
    """Aggregate a student's quiz completion + score across every chapter of a
    course. A quiz only counts toward totals once it has at least one question
    (an empty quiz shell isn't something a student can be "missing")."""
    chapters = conn.execute(
        "SELECT id, title FROM chapters WHERE course_id=? ORDER BY sort_order, id", (course_id,)
    ).fetchall()
    quiz_list = []
    sum_score, sum_max = 0.0, 0.0
    for ch in chapters:
        quizzes = conn.execute("SELECT id, title FROM quizzes WHERE chapter_id=? ORDER BY id", (ch["id"],)).fetchall()
        for qz in quizzes:
            questions = conn.execute("SELECT id, points FROM questions WHERE quiz_id=?", (qz["id"],)).fetchall()
            n_q = len(questions)
            max_score = sum(q["points"] for q in questions)
            subs = {}
            if n_q:
                qids = [q["id"] for q in questions]
                placeholders = ",".join("?" * len(qids))
                rows = conn.execute(
                    f"SELECT * FROM submissions WHERE student_id=? AND question_id IN ({placeholders})",
                    [student_id] + qids,
                ).fetchall()
                subs = {r["question_id"]: r for r in rows}
            answered = len(subs)
            pending = sum(1 for r in subs.values() if r["grading_state"] == "pending")
            score = sum((r["score"] or 0) for r in subs.values() if r["score"] is not None)
            if n_q == 0:
                status = "empty"
            elif answered == 0:
                status = "not_started"
            elif answered < n_q:
                status = "in_progress"
            elif pending > 0:
                status = "pending_grading"
            else:
                status = "completed"
            if n_q > 0:
                sum_score += score
                sum_max += max_score
            quiz_list.append({
                "quiz_id": qz["id"], "quiz_title": qz["title"], "chapter_title": ch["title"],
                "status": status, "score": round(score, 2), "max_score": round(max_score, 2),
                "answered": answered, "total_questions": n_q,
            })
    gradable = [q for q in quiz_list if q["total_questions"] > 0]
    total = len(gradable)
    completed = sum(1 for q in gradable if q["status"] == "completed")
    percent_complete = round(completed / total * 100, 1) if total else 0.0
    percent_score = round(sum_score / sum_max * 100, 1) if sum_max else None
    missing = [q for q in gradable if q["status"] != "completed"]
    return {
        "quizzes": quiz_list,
        "total_quizzes": total,
        "completed_quizzes": completed,
        "percent_complete": percent_complete,
        "total_score": round(sum_score, 2),
        "total_max": round(sum_max, 2),
        "percent_score": percent_score,
        "grade": letter_grade(percent_score),
        "missing_quizzes": missing,
    }


# ---- Instructor: chapters ----
@app.get("/api/instructor/courses/{course_id}/chapters")
def instructor_list_chapters(course_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    rows = conn.execute(
        "SELECT * FROM chapters WHERE course_id=? ORDER BY sort_order, id", (course_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/instructor/courses/{course_id}/chapters")
def instructor_create_chapter(course_id: int, body: ChapterBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    if not body.title.strip():
        raise HTTPException(400, "กรุณาระบุชื่อบท")
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    cur = conn.execute(
        "INSERT INTO chapters (course_id,title,sort_order) VALUES (?,?,?)",
        (course_id, body.title.strip(), body.sort_order or 0),
    )
    conn.commit()
    conn.close()
    return {"id": cur.lastrowid}


@app.patch("/api/instructor/chapters/{chapter_id}")
def instructor_update_chapter(chapter_id: int, body: ChapterBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    conn.execute(
        "UPDATE chapters SET title=?, sort_order=? WHERE id=?",
        (body.title.strip(), body.sort_order or 0, chapter_id),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/instructor/chapters/{chapter_id}")
def instructor_delete_chapter(chapter_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    quiz_ids = [r["id"] for r in conn.execute("SELECT id FROM quizzes WHERE chapter_id=?", (chapter_id,)).fetchall()]
    for qid in quiz_ids:
        q_ids = [r["id"] for r in conn.execute("SELECT id FROM questions WHERE quiz_id=?", (qid,)).fetchall()]
        for qqid in q_ids:
            conn.execute("DELETE FROM submissions WHERE question_id=?", (qqid,))
        conn.execute("DELETE FROM questions WHERE quiz_id=?", (qid,))
    conn.execute("DELETE FROM quizzes WHERE chapter_id=?", (chapter_id,))
    conn.execute("DELETE FROM materials WHERE chapter_id=?", (chapter_id,))
    conn.execute("DELETE FROM chapters WHERE id=?", (chapter_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---- Instructor: materials ----
@app.get("/api/instructor/chapters/{chapter_id}/materials")
def instructor_list_materials(chapter_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    rows = conn.execute(
        "SELECT * FROM materials WHERE chapter_id=? ORDER BY sort_order, id", (chapter_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/instructor/chapters/{chapter_id}/materials")
def instructor_create_material(chapter_id: int, body: MaterialBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    if body.type not in ("text", "image", "audio"):
        raise HTTPException(400, "type ต้องเป็น text, image หรือ audio")
    if not body.title.strip():
        raise HTTPException(400, "กรุณาระบุชื่อเนื้อหา")
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    cur = conn.execute(
        "INSERT INTO materials (chapter_id,type,title,body,file_url,sort_order) VALUES (?,?,?,?,?,?)",
        (chapter_id, body.type, body.title.strip(), body.body, body.file_url, body.sort_order or 0),
    )
    conn.commit()
    conn.close()
    return {"id": cur.lastrowid}


@app.delete("/api/instructor/materials/{material_id}")
def instructor_delete_material(material_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    row = conn.execute(
        """SELECT m.*, c.instructor_id FROM materials m
           JOIN chapters ch ON ch.id = m.chapter_id
           JOIN courses c ON c.id = ch.course_id WHERE m.id=?""",
        (material_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบเนื้อหานี้")
    if row["instructor_id"] != instr["id"]:
        raise HTTPException(403, "not your material")
    conn.execute("DELETE FROM materials WHERE id=?", (material_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---- Instructor: quizzes & questions ----
@app.get("/api/instructor/chapters/{chapter_id}/quizzes")
def instructor_list_quizzes(chapter_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    rows = conn.execute("SELECT * FROM quizzes WHERE chapter_id=? ORDER BY id", (chapter_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/instructor/chapters/{chapter_id}/quizzes")
def instructor_create_quiz(chapter_id: int, body: QuizBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    if not body.title.strip():
        raise HTTPException(400, "กรุณาระบุชื่อควิซ")
    conn = db()
    _owned_chapter(conn, chapter_id, instr["id"])
    cur = conn.execute(
        "INSERT INTO quizzes (chapter_id,title,instructions) VALUES (?,?,?)",
        (chapter_id, body.title.strip(), body.instructions),
    )
    conn.commit()
    conn.close()
    return {"id": cur.lastrowid}


@app.delete("/api/instructor/quizzes/{quiz_id}")
def instructor_delete_quiz(quiz_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_quiz(conn, quiz_id, instr["id"])
    q_ids = [r["id"] for r in conn.execute("SELECT id FROM questions WHERE quiz_id=?", (quiz_id,)).fetchall()]
    for qid in q_ids:
        conn.execute("DELETE FROM submissions WHERE question_id=?", (qid,))
    conn.execute("DELETE FROM questions WHERE quiz_id=?", (quiz_id,))
    conn.execute("DELETE FROM quizzes WHERE id=?", (quiz_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/instructor/quizzes/{quiz_id}/questions")
def instructor_list_questions(quiz_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_quiz(conn, quiz_id, instr["id"])
    rows = conn.execute("SELECT * FROM questions WHERE quiz_id=? ORDER BY sort_order, id", (quiz_id,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        d["params"] = json.loads(d.pop("params_json"))
        out.append(d)
    return out


@app.post("/api/instructor/quizzes/{quiz_id}/questions")
def instructor_create_question(quiz_id: int, body: QuestionBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    if body.quiz_type not in QUIZ_TYPES:
        raise HTTPException(400, f"quiz_type ต้องเป็นหนึ่งใน {QUIZ_TYPES}")
    if not body.prompt.strip():
        raise HTTPException(400, "กรุณาระบุคำถาม")
    conn = db()
    _owned_quiz(conn, quiz_id, instr["id"])
    cur = conn.execute(
        "INSERT INTO questions (quiz_id,quiz_type,prompt,points,sort_order,params_json) VALUES (?,?,?,?,?,?)",
        (quiz_id, body.quiz_type, body.prompt.strip(), body.points or 1, body.sort_order or 0, json.dumps(body.params)),
    )
    conn.commit()
    conn.close()
    return {"id": cur.lastrowid}


@app.delete("/api/instructor/questions/{question_id}")
def instructor_delete_question(question_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    row = conn.execute(
        """SELECT q.*, c.instructor_id FROM questions q
           JOIN quizzes quiz ON quiz.id = q.quiz_id
           JOIN chapters ch ON ch.id = quiz.chapter_id
           JOIN courses c ON c.id = ch.course_id WHERE q.id=?""",
        (question_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบคำถามนี้")
    if row["instructor_id"] != instr["id"]:
        raise HTTPException(403, "not your question")
    conn.execute("DELETE FROM submissions WHERE question_id=?", (question_id,))
    conn.execute("DELETE FROM questions WHERE id=?", (question_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---- Instructor: grading & scores ----
@app.get("/api/instructor/quizzes/{quiz_id}/submissions")
def instructor_quiz_submissions(quiz_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    quiz = _owned_quiz(conn, quiz_id, instr["id"])
    questions = conn.execute("SELECT * FROM questions WHERE quiz_id=? ORDER BY sort_order, id", (quiz_id,)).fetchall()
    max_score = sum(q["points"] for q in questions)
    students = conn.execute(
        """SELECT u.id, u.full_name, u.student_id FROM enrollments e
           JOIN users u ON u.id = e.user_id WHERE e.course_id=? ORDER BY u.full_name""",
        (quiz["course_id"],),
    ).fetchall()
    result = []
    for s in students:
        subs_by_qid = {}
        if questions:
            placeholders = ",".join("?" * len(questions))
            subs = conn.execute(
                f"SELECT * FROM submissions WHERE student_id=? AND question_id IN ({placeholders})",
                [s["id"]] + [q["id"] for q in questions],
            ).fetchall()
            subs_by_qid = {r["question_id"]: r for r in subs}
        answers, total = [], 0
        for q in questions:
            sub = subs_by_qid.get(q["id"])
            answers.append({
                "question_id": q["id"], "quiz_type": q["quiz_type"], "prompt": q["prompt"], "points": q["points"],
                "answer": json.loads(sub["answer_json"]) if sub else None,
                "is_correct": bool(sub["is_correct"]) if sub and sub["is_correct"] is not None else None,
                "score": sub["score"] if sub else None,
                "grading_state": sub["grading_state"] if sub else None,
                "feedback": sub["feedback"] if sub else None,
                "submission_id": sub["id"] if sub else None,
            })
            if sub and sub["score"] is not None:
                total += sub["score"]
        result.append({
            "student_id": s["id"], "full_name": s["full_name"], "student_no": s["student_id"],
            "total_score": total, "max_score": max_score, "answers": answers,
        })
    conn.close()
    return result


@app.patch("/api/instructor/submissions/{submission_id}")
def instructor_grade_submission(submission_id: int, body: GradeSubmissionBody, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    row = conn.execute(
        """SELECT s.*, c.instructor_id FROM submissions s
           JOIN questions q ON q.id = s.question_id
           JOIN quizzes quiz ON quiz.id = q.quiz_id
           JOIN chapters ch ON ch.id = quiz.chapter_id
           JOIN courses c ON c.id = ch.course_id WHERE s.id=?""",
        (submission_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "ไม่พบคำตอบนี้")
    if row["instructor_id"] != instr["id"]:
        raise HTTPException(403, "not your submission")
    conn.execute(
        "UPDATE submissions SET score=?, feedback=?, grading_state='graded', is_correct=? WHERE id=?",
        (body.score, body.feedback, 1 if body.score > 0 else 0, submission_id),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


# ================= STUDENT: chapters / materials / quizzes =================
@app.get("/api/student/courses/{course_id}/chapters")
def student_list_chapters(course_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    enrolled = conn.execute(
        "SELECT 1 FROM enrollments WHERE user_id=? AND course_id=?", (student["id"], course_id)
    ).fetchone()
    if not enrolled:
        conn.close()
        raise HTTPException(403, "คุณไม่มีสิทธิ์เข้าคอร์สนี้")
    rows = conn.execute("SELECT * FROM chapters WHERE course_id=? ORDER BY sort_order, id", (course_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/student/chapters/{chapter_id}/materials")
def student_list_materials(chapter_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    _student_owned_chapter(conn, chapter_id, student["id"])
    rows = conn.execute("SELECT * FROM materials WHERE chapter_id=? ORDER BY sort_order, id", (chapter_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/student/chapters/{chapter_id}/quizzes")
def student_list_quizzes(chapter_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    _student_owned_chapter(conn, chapter_id, student["id"])
    rows = conn.execute("SELECT * FROM quizzes WHERE chapter_id=? ORDER BY id", (chapter_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/student/quizzes/{quiz_id}/questions")
def student_list_questions(quiz_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    _student_owned_quiz(conn, quiz_id, student["id"])
    rows = conn.execute("SELECT * FROM questions WHERE quiz_id=? ORDER BY sort_order, id", (quiz_id,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        params = json.loads(r["params_json"])
        out.append({
            "id": r["id"], "quiz_type": r["quiz_type"], "prompt": r["prompt"], "points": r["points"],
            "params": strip_answer_key(r["quiz_type"], params),
        })
    return out


@app.post("/api/student/questions/{question_id}/submit")
def student_submit_answer(question_id: int, body: SubmitAnswerBody, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    q = conn.execute(
        """SELECT q.*, ch.course_id FROM questions q
           JOIN quizzes quiz ON quiz.id = q.quiz_id
           JOIN chapters ch ON ch.id = quiz.chapter_id WHERE q.id=?""",
        (question_id,),
    ).fetchone()
    if not q:
        conn.close()
        raise HTTPException(404, "ไม่พบคำถามนี้")
    enrolled = conn.execute(
        "SELECT 1 FROM enrollments WHERE user_id=? AND course_id=?", (student["id"], q["course_id"])
    ).fetchone()
    if not enrolled:
        conn.close()
        raise HTTPException(403, "คุณไม่มีสิทธิ์ทำควิซนี้")
    params = json.loads(q["params_json"])
    correct = grade_answer(q["quiz_type"], params, body.answer)
    if q["quiz_type"] == "file_upload":
        is_correct, score, grading_state = None, None, "pending"
    else:
        is_correct = 1 if correct else 0
        score = q["points"] if correct else 0
        grading_state = "auto"
    conn.execute(
        """INSERT INTO submissions (question_id,student_id,answer_json,is_correct,score,grading_state)
           VALUES (?,?,?,?,?,?)
           ON CONFLICT(question_id,student_id) DO UPDATE SET
             answer_json=excluded.answer_json, is_correct=excluded.is_correct,
             score=excluded.score, grading_state=excluded.grading_state, submitted_at=datetime('now')""",
        (question_id, student["id"], json.dumps(body.answer), is_correct, score, grading_state),
    )
    conn.commit()
    conn.close()
    return {"is_correct": bool(is_correct) if is_correct is not None else None, "grading_state": grading_state, "score": score}


@app.get("/api/student/quizzes/{quiz_id}/result")
def student_quiz_result(quiz_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    _student_owned_quiz(conn, quiz_id, student["id"])
    questions = conn.execute("SELECT * FROM questions WHERE quiz_id=? ORDER BY sort_order, id", (quiz_id,)).fetchall()
    max_score = sum(q["points"] for q in questions)
    items, total = [], 0
    for q in questions:
        sub = conn.execute(
            "SELECT * FROM submissions WHERE question_id=? AND student_id=?", (q["id"], student["id"])
        ).fetchone()
        items.append({
            "question_id": q["id"], "quiz_type": q["quiz_type"], "prompt": q["prompt"], "points": q["points"],
            "is_correct": bool(sub["is_correct"]) if sub and sub["is_correct"] is not None else None,
            "score": sub["score"] if sub else None,
            "grading_state": sub["grading_state"] if sub else None,
            "feedback": sub["feedback"] if sub else None,
        })
        if sub and sub["score"] is not None:
            total += sub["score"]
    conn.close()
    return {"total_score": total, "max_score": max_score, "items": items}


@app.get("/api/student/courses/{course_id}/progress")
def student_progress(course_id: int, student: sqlite3.Row = Depends(require_role("student"))):
    conn = db()
    enrolled = conn.execute(
        "SELECT 1 FROM enrollments WHERE user_id=? AND course_id=?", (student["id"], course_id)
    ).fetchone()
    if not enrolled:
        conn.close()
        raise HTTPException(403, "คุณไม่มีสิทธิ์เข้าคอร์สนี้")
    result = compute_progress(conn, student["id"], course_id)
    conn.close()
    return result


@app.get("/api/instructor/courses/{course_id}/grades")
def instructor_grades(course_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    _owned_course(conn, course_id, instr["id"])
    roster = conn.execute(
        "SELECT u.* FROM enrollments e JOIN users u ON u.id=e.user_id WHERE e.course_id=? ORDER BY u.full_name",
        (course_id,),
    ).fetchall()
    out = []
    for u in roster:
        prog = compute_progress(conn, u["id"], course_id)
        out.append({
            "student_id": u["id"], "full_name": u["full_name"], "student_code": u["student_id"],
            "percent_complete": prog["percent_complete"], "completed_quizzes": prog["completed_quizzes"],
            "total_quizzes": prog["total_quizzes"], "percent_score": prog["percent_score"],
            "grade": prog["grade"], "total_score": prog["total_score"], "total_max": prog["total_max"],
            "missing_quizzes": prog["missing_quizzes"],
        })
    conn.close()
    return out


@app.get("/api/instructor/courses/{course_id}/grades/export")
def instructor_export_grades(course_id: int, instr: sqlite3.Row = Depends(require_role("instructor"))):
    conn = db()
    course = _owned_course(conn, course_id, instr["id"])
    roster = conn.execute(
        "SELECT u.* FROM enrollments e JOIN users u ON u.id=e.user_id WHERE e.course_id=? ORDER BY u.full_name",
        (course_id,),
    ).fetchall()
    quiz_rows = conn.execute(
        """SELECT qz.id, qz.title, ch.title AS chapter_title FROM quizzes qz
           JOIN chapters ch ON ch.id = qz.chapter_id WHERE ch.course_id=?
           ORDER BY ch.sort_order, ch.id, qz.id""",
        (course_id,),
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "คะแนน"
    header = ["ชื่อ-นามสกุล", "รหัสนักศึกษา"] + [f"{q['chapter_title']} - {q['title']}" for q in quiz_rows] + [
        "คะแนนรวม", "คะแนนเต็ม", "เปอร์เซ็นต์ (%)", "เกรด", "ความคืบหน้า (%)",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for u in roster:
        prog = compute_progress(conn, u["id"], course_id)
        qmap = {q["quiz_id"]: q for q in prog["quizzes"]}
        row = [u["full_name"], u["student_id"] or "-"]
        for qz in quiz_rows:
            info = qmap.get(qz["id"])
            if not info or info["total_questions"] == 0:
                row.append("-")
            elif info["status"] == "not_started":
                row.append("ยังไม่ได้ทำ")
            elif info["status"] in ("in_progress", "pending_grading"):
                row.append(f"{info['score']}/{info['max_score']} (รอตรวจ)")
            else:
                row.append(f"{info['score']}/{info['max_score']}")
        row += [
            prog["total_score"], prog["total_max"],
            prog["percent_score"] if prog["percent_score"] is not None else "-",
            prog["grade"] or "-", prog["percent_complete"],
        ]
        ws.append(row)
    conn.close()

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 40)
    ws.freeze_panes = "C2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"grades_{course['slug']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/dashboard")
def admin_dashboard(admin: sqlite3.Row = Depends(require_role("admin"))):
    conn = db()
    courses = conn.execute("SELECT * FROM courses ORDER BY title").fetchall()
    out = []
    for c in courses:
        roster = conn.execute(
            "SELECT u.* FROM enrollments e JOIN users u ON u.id=e.user_id WHERE e.course_id=? ORDER BY u.full_name",
            (c["id"],),
        ).fetchall()
        students = []
        for u in roster:
            prog = compute_progress(conn, u["id"], c["id"])
            students.append({
                "student_id": u["id"], "full_name": u["full_name"], "student_code": u["student_id"],
                "percent_complete": prog["percent_complete"], "percent_score": prog["percent_score"],
                "grade": prog["grade"],
            })
        scored = [s["percent_score"] for s in students if s["percent_score"] is not None]
        out.append({
            "course_id": c["id"], "course_title": c["title"], "students": students,
            "student_count": len(students),
            "avg_complete": round(sum(s["percent_complete"] for s in students) / len(students), 1) if students else 0,
            "avg_score": round(sum(scored) / len(scored), 1) if scored else None,
        })
    conn.close()
    return out


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8010)
